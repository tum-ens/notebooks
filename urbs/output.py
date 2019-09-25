import pandas as pd
from .input import get_input
from .pyomoio import get_entity, get_entities
from .util import is_string
from openpyxl import load_workbook
import os
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
import psutil
import time


class TerminalAndFileWriter(object):
    """
    This class allows to write to the Terminal and a file at the same time. It is used for the option save_terminal_output run_scenario_decomposition in runme.py.
    """
    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()

    def flush(self) :
        for f in self.files:
            f.flush()


def get_constants(instance):
    """Return summary DataFrames for important variables

    Usage:
        costs, cpro, ctra, csto = get_constants(instance)

    Args:
        instance: a urbs model instance

    Returns:
        (costs, cpro, ctra, csto) tuple

    Example:
        >>> import pyomo.environ
        >>> from pyomo.opt.base import SolverFactory
        >>> data = read_excel('mimo-example.xlsx')
        >>> prob = create_model(data, range(1,25))
        >>> optim = SolverFactory('glpk')
        >>> result = optim.solve(prob)
        >>> cap_pro = get_constants(prob)[1]['Total']
        >>> cap_pro.xs('Wind park', level='Process').apply(int)
        Site
        Mid      13000
        North    23258
        South        0
        Name: Total, dtype: int64
    """
    costs = get_entity(instance, 'costs')
    cpro = get_entities(instance, ['cap_pro', 'cap_pro_new'])
    ctra = get_entities(instance, ['cap_tra', 'cap_tra_new'])
    csto = get_entities(instance, ['cap_sto_c', 'cap_sto_c_new',
                                   'cap_sto_p', 'cap_sto_p_new'])

    # better labels and index names and return sorted
    if not cpro.empty:
        cpro.index.names = ['Site', 'Process']
        cpro.columns = ['Total', 'New']
        cpro.sort_index(inplace=True)
    if not ctra.empty:
        ctra.index.names = ['Site In', 'Site Out', 'Transmission', 'Commodity']
        ctra.columns = ['Total', 'New']
        ctra.sort_index(inplace=True)
    if not csto.empty:
        csto.columns = ['C Total', 'C New', 'P Total', 'P New']
        csto.sort_index(inplace=True)

    return costs, cpro, ctra, csto


def get_timeseries(instance, com, sites, timesteps=None):
    """Return DataFrames of all timeseries referring to given commodity

    Usage:
        (created, consumed, stored, imported, exported,
         dsm) = get_timeseries(instance, commodity, sites, timesteps)

    Args:
        instance: a urbs model instance
        com: a commodity name
        sites: a site name or list of site names
        timesteps: optional list of timesteps, default: all modelled timesteps

    Returns:
        a tuple of (created, consumed, storage, imported, exported, dsm) with
        DataFrames timeseries. These are:

        - created: timeseries of commodity creation, including stock source
        - consumed: timeseries of commodity consumption, including demand
        - storage: timeseries of commodity storage (level, stored, retrieved)
        - imported: timeseries of commodity import
        - exported: timeseries of commodity export
        - dsm: timeseries of demand-side management
    """
    if timesteps is None:
        # default to all simulated timesteps
        timesteps = sorted(get_entity(instance, 'tm').index)
    else:
        timesteps = sorted(timesteps)  # implicit: convert range to list

    if is_string(sites):
        # wrap single site name into list
        sites = [sites]

    # DEMAND
    # default to zeros if commodity has no demand, get timeseries
    try:
        # select relevant timesteps (=rows)
        # select commodity (xs), then the sites from remaining simple columns
        # and sum all together to form a Series
        demand = (get_input(instance, 'demand').loc[timesteps]
                                               .xs(com, axis=1, level=1)[sites]
                                               .sum(axis=1))
    except KeyError:
        demand = pd.Series(0, index=timesteps)
    demand.name = 'Demand'

    # STOCK
    eco = get_entity(instance, 'e_co_stock')
    eco = eco.xs([com, 'Stock'], level=['com', 'com_type'])
    try:
        stock = eco.unstack()[sites].sum(axis=1)
    except KeyError:
        stock = pd.Series(0, index=timesteps)
    stock.name = 'Stock'

    # PROCESS
    created = get_entity(instance, 'e_pro_out')
    created = created.xs(com, level='com').loc[timesteps]
    try:
        created = created.unstack(level='sit')[sites].fillna(0).sum(axis=1)
        created = created.unstack(level='pro')
        created = drop_all_zero_columns(created)
    except KeyError:
        created = pd.DataFrame(index=timesteps)

    consumed = get_entity(instance, 'e_pro_in')
    consumed = consumed.xs(com, level='com').loc[timesteps]
    try:
        consumed = consumed.unstack(level='sit')[sites].fillna(0).sum(axis=1)
        consumed = consumed.unstack(level='pro')
        consumed = drop_all_zero_columns(consumed)
    except KeyError:
        consumed = pd.DataFrame(index=timesteps)

    # TRANSMISSION
    other_sites = get_input(instance, 'site').index.difference(sites)

    # if commodity is transportable
    df_transmission = get_input(instance, 'transmission')
    if com in set(df_transmission.index.get_level_values('Commodity')):
        imported = get_entity(instance, 'e_tra_out')
        imported = imported.loc[timesteps].xs(com, level='com')
        imported = imported.unstack(level='tra').sum(axis=1)
        imported = imported.unstack(level='sit_')[sites].fillna(0).sum(axis=1)
        imported = imported.unstack(level='sit')

        internal_import = imported[sites].sum(axis=1)  # ...from sites
        imported = imported[other_sites]  # ...from other_sites
        imported = drop_all_zero_columns(imported)

        exported = get_entity(instance, 'e_tra_in')
        exported = exported.loc[timesteps].xs(com, level='com')
        exported = exported.unstack(level='tra').sum(axis=1)
        exported = exported.unstack(level='sit')[sites].fillna(0).sum(axis=1)
        exported = exported.unstack(level='sit_')

        internal_export = exported[sites].sum(axis=1)  # ...to sites (internal)
        exported = exported[other_sites]  # ...to other_sites
        exported = drop_all_zero_columns(exported)
    else:
        imported = pd.DataFrame(index=timesteps)
        exported = pd.DataFrame(index=timesteps)
        internal_export = pd.Series(0, index=timesteps)
        internal_import = pd.Series(0, index=timesteps)

    # to be discussed: increase demand by internal transmission losses
    internal_transmission_losses = internal_export - internal_import
    demand = demand + internal_transmission_losses

    # STORAGE
    # group storage energies by commodity
    # select all entries with desired commodity co
    stored = get_entities(instance, ['e_sto_con', 'e_sto_in', 'e_sto_out'])
    try:
        stored = stored.loc[timesteps].xs(com, level='com')
        stored = stored.groupby(level=['t', 'sit']).sum()
        stored = stored.loc[(slice(None), sites), :].sum(level='t')
        stored.columns = ['Level', 'Stored', 'Retrieved']
    except (KeyError, ValueError):
        stored = pd.DataFrame(0, index=timesteps,
                              columns=['Level', 'Stored', 'Retrieved'])

    demand.name = 'Unshifted'

    # JOINS
    created = created.join(stock)  # show stock as created
    consumed = consumed.join(demand.rename('Demand'))

    return created, consumed, stored, imported, exported


def drop_all_zero_columns(df):
    """ Drop columns from DataFrame if they contain only zeros.

    Args:
        df: a DataFrame

    Returns:
        the DataFrame without columns that only contain zeros
    """
    return df.loc[:, (df != 0).any(axis=0)]


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow , index=False, **to_excel_kwargs)

    else:
        df.to_excel(writer, sheet_name, startrow=startrow, header=None, index=False)

    # save the workbook
    writer.save()


def prepare_result_directory(result_name):
    """ create a time stamped directory within the result folder """
    # timestamp for result directory
    now = datetime.now().strftime('%Y%m%dT%H%M')

    # create result directory if not existent
    result_dir = os.path.join('result', '{}-{}'.format(result_name, now))
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)

    return result_dir


def plot_convergence(iterations, plot_lower_bounds, plot_upper_bounds,  result_dir, sce, run_normal=False, normal=None):
    """Show the convergence plot of the bounds for the different run_scenario methods (run_benders, run_regional, run_sddp)

    Args:
        iterations: list of iterations
        plot_lower_bounds: list of the lower bounds corresponding to each iteration
        plot_upper_bounds: list of the upper bounds corresponding to each iteration
        normal: list of normals corresponding to each iteration
        result_dir: directory, where the plot is saved
        sce: scenario name

    Returns: Nothing
    """
    iterator = np.array(iterations)
    lb = np.array(plot_lower_bounds)
    if run_normal:
        opt = np.array(normal)
    ub = np.array(plot_upper_bounds)
    fig = plt.figure()

    plt.plot(iterator, lb)#
    if run_normal:
        plt.plot(iterator, opt)
    plt.plot(iterator, ub)
    # set logarithmic axes, title and grid
    plt.yscale('log')
    plt.xlabel('Iterations')
    plt.ylabel('Objective')
    plt.title('Converging bounds')
    plt.grid(True)
    # plt.show()

    fig.savefig(os.path.join(result_dir, 'bounds-{}.png'.format(sce)), dpi=300)


def create_benders_output_table(print_omega=False):
    """Output information for every iteration in the benders loop of run_benders and run_regional.
    This function outputs the original problem objective and the names of the columns.

    Args:
        print_omega: if true omega is displayed in the table
    """
    print()
    print('   i',end='')
    if print_omega: print("{:>10}".format('omega'),end='')
    print(
          "{:>14}".format('Master Eta'),
          "{:>14}".format('Sub Lambda'),
          "{:>14}".format('Lower Bound'),
          "{:>14}".format('Upper Bound'),
          "{:>14}".format('Dual gap'),
          "{:>14}".format('Master obj')
          )


def create_benders_output_table_sddp():
    """Output information for every iteration in the benders loop of run_sddp.
    This function writes the names of the columns.
    """
    print()
    print('   i',
          "{:>14}".format('Master Eta'),
          "{:>14}".format('LB'),
          "{:>14}".format('UB (latest)'),
          "{:>14}".format('UB (last 10)'),
          "{:>14}".format('stddev'),
          "{:>14}".format('Dual gap'),
          "{:>16}".format('Master obj')
          )


def update_benders_output_table(i, master, master_eta, sub, lower_bound, upper_bound, gap, print_omega=False):
    """Complete output for one iteration of the benders loop of run_benders, run_regional.

    Args:
        i: iteration number
        master: master problem instance
        master_eta:
        sub: list of sub problem instances
        lower_bound: lower bound
        upper_bound: upper bound
        gap: dual gap
        print_omega: if true omega is displayed in the table
    """
    print('{:4}'.format(i), '   ',end='')
    if print_omega: print('{:4}'.format(sum(sub[str(inst)].omega() for inst in sub)),'    ',end='')
    print(
          '{:10.3e}'.format(master_eta), '   ',
          '{:10.3e}'.format(sum(sub[inst].Lambda() for inst in sub)), '   ',
          '{:10.3e}'.format(lower_bound), '   ',
          '{:10.3e}'.format(upper_bound), '   ',
          '{:10.3e}'.format(gap), '   ',
          '{:12.5e}'.format(master.obj()))


def update_benders_output_table_sddp(i, master, lower_bound, upper_bound, avg, stddev, gap, master_objective):
    """Complete output for one iteration of the benders loop of run_sddp.

    Args:
        i: iteration number
        lower_bound: lower bound
        upper_bound: upper bound
        avg: average of the last ten upper bounds
        stddev: standard deviation within the last ten upper bounds
        gap: dual gap
        master_objective: objective of the master problem
    """
    print('{:4}'.format(i), '   ',
          '{:10.3e}'.format(master.eta()), '   ',
          '{:10.3e}'.format(lower_bound), '   ',
          '{:10.3e}'.format(upper_bound), '   ',
          '{:10.3e}'.format(avg), '   ',
          '{:10.3e}'.format(stddev), '   ',
          '{:10.3e}'.format(gap), '   ',
          '{:12.5e}'.format(master_objective))


def create_tracking_file(track_file, start_time):
    """
    Creates a file to track hardware usage

    Args:
        track_file: file, in which to save the tracking
        start_time: time at which the scenario decomposition was started

    Returns: The process for tracking, it is needed later to update the tracking file
    """
    process = psutil.Process(os.getpid())
    process.cpu_percent()
    create_time = time.time() - start_time
    # save hardware usage
    solve_time_orig = time.time() - start_time
    with open(track_file, 'w') as help_file:
        print('Create time: ', create_time, '\r\n', file=help_file)
        print('   i       Memory       CPU Time       CPU Percentage       Time  \r\n', file=help_file)
        print('original',
              '{:10.3e}'.format(process.memory_info()[0]), '   ',
              '{:10.3e}'.format(process.cpu_times()[0]), '   ',
              process.cpu_percent(), '            ',
              solve_time_orig,
              '\r\n', file=help_file)
    return process


def update_tracking_file(track_file,i,start_time, process):
    """
    Update the tracking file

    Args:
        track_file: file, in which to save the tracking
        i: iteration number
        start_time: time at which the scenario decomposition was started
        process: process created by create_tracking_file for tracking
    """
    with open(track_file, 'a') as help_file:
        print('{:4}'.format(i), '   ',
              '{:10.3e}'.format(process.memory_info()[0]), '   ',
              '{:10.3e}'.format(process.cpu_times()[0]), '   ',
              '{:10.3e}'.format(process.cpu_percent()), '         ',
              time.time() - start_time,
              '\r\n',
              file=help_file)


